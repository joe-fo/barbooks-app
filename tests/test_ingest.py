"""Tests for the ingest.py source data fetcher and validator."""

from __future__ import annotations

from typing import Optional

import openpyxl
import pytest
from bs4 import BeautifulSoup

from app.domain.models import Page
from ingest import (
    _clean_name,
    _extract_answer_count,
    _extract_title,
    _find_xlsx,
    _parse_ordered_list_items,
    _parse_table_items,
    _write_page_to_spreadsheet,
    parse_page_data,
)

# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------


def _html(body: str) -> str:
    return (
        f"<html><head><title>Test Page - ESPN</title></head><body>{body}</body></html>"
    )


def _soup(html: str) -> BeautifulSoup:
    return BeautifulSoup(html, "html.parser")


# ---------------------------------------------------------------------------
# _clean_name
# ---------------------------------------------------------------------------


class TestCleanName:
    def test_strips_abbreviated_suffix(self):
        assert _clean_name("Eli ManningE. Manning") == "Eli Manning"

    def test_strips_abbreviated_suffix_different_player(self):
        assert _clean_name("Dak PrescottD. Prescott") == "Dak Prescott"

    def test_leaves_clean_name_unchanged(self):
        assert _clean_name("Jerry Rice") == "Jerry Rice"

    def test_leaves_multiword_name_unchanged(self):
        assert _clean_name("LaDainian Tomlinson") == "LaDainian Tomlinson"

    def test_leaves_empty_string_unchanged(self):
        assert _clean_name("") == ""


# ---------------------------------------------------------------------------
# _extract_title
# ---------------------------------------------------------------------------


class TestExtractAnswerCount:
    def test_top_10(self):
        assert _extract_answer_count("Top 10 Career Passing Yard Leaders") == 10

    def test_top_5(self):
        assert _extract_answer_count("Top 5 NFL Rushers") == 5

    def test_case_insensitive(self):
        assert _extract_answer_count("top 25 all-time scorers") == 25

    def test_no_top_n_returns_zero(self):
        assert _extract_answer_count("NFL All-Time Touchdown Leaders") == 0

    def test_empty_string_returns_zero(self):
        assert _extract_answer_count("") == 0


class TestExtractTitle:
    def test_prefers_h1(self):
        soup = _soup(
            "<html><head><title>Site</title></head><body><h1>My List</h1></body></html>"
        )
        assert _extract_title(soup) == "My List"

    def test_falls_back_to_title(self):
        soup = _soup(
            "<html><head><title>Page Title - ESPN</title></head><body></body></html>"
        )
        assert _extract_title(soup) == "Page Title"

    def test_returns_empty_when_no_title(self):
        soup = _soup("<html><body></body></html>")
        assert _extract_title(soup) == ""

    def test_strips_pipe_suffix(self):
        soup = _soup(
            "<html><head><title>Best Title | ESPN</title></head><body></body></html>"
        )
        assert _extract_title(soup) == "Best Title"


# ---------------------------------------------------------------------------
# _parse_table_items
# ---------------------------------------------------------------------------


class TestParseTableItems:
    def _table_html(self, headers: list[str], rows: list[list[str]]) -> str:
        header_row = "<tr>" + "".join(f"<th>{h}</th>" for h in headers) + "</tr>"
        body_rows = "".join(
            "<tr>" + "".join(f"<td>{c}</td>" for c in row) + "</tr>" for row in rows
        )
        return f"<table>{header_row}{body_rows}</table>"

    def test_basic_ranked_table(self):
        html = self._table_html(
            ["Rank", "Player", "TDs"],
            [["1", "Jerry Rice", "208"], ["2", "Emmitt Smith", "175"]],
        )
        soup = _soup(html)
        table = soup.find("table")
        items, stat_label = _parse_table_items(table)
        assert len(items) == 2
        assert items[0].name == "Jerry Rice"
        assert items[0].rank == 1
        assert items[0].stat_value == "208"
        assert "TD" in stat_label.lower() or stat_label == "TDs"

    def test_returns_empty_for_empty_table(self):
        soup = _soup("<table></table>")
        table = soup.find("table")
        items, stat_label = _parse_table_items(table)
        assert items == []
        assert stat_label == ""

    def test_skips_rows_with_no_name(self):
        html = self._table_html(
            ["Rank", "Player", "Yards"],
            [["1", "", "1846"], ["2", "Barry Sanders", "2358"]],
        )
        soup = _soup(html)
        table = soup.find("table")
        items, _ = _parse_table_items(table)
        assert len(items) == 1
        assert items[0].name == "Barry Sanders"

    def test_handles_non_numeric_rank(self):
        html = self._table_html(
            ["Rank", "Player", "TDs"],
            [["T1", "Jerry Rice", "208"]],
        )
        soup = _soup(html)
        table = soup.find("table")
        items, _ = _parse_table_items(table)
        assert len(items) == 1
        assert items[0].rank == 1  # parsed digit from "T1"

    def test_strips_abbreviated_name_suffix(self):
        """Name cells that embed full + abbreviated form must not produce duplicates."""
        # Simulate a table cell like <td>Eli Manning<abbr>E. Manning</abbr></td>
        # get_text(strip=True) produces "Eli ManningE. Manning" without the fix.
        html = (
            "<table>"
            "<tr><th>Rank</th><th>Player</th><th>TDs</th></tr>"
            "<tr><td>1</td>"
            "<td>Eli Manning<abbr>E. Manning</abbr></td><td>117</td></tr>"
            "<tr><td>2</td>"
            "<td>Dak Prescott<abbr>D. Prescott</abbr></td><td>98</td></tr>"
            "<tr><td>3</td><td>Tom Brady<abbr>T. Brady</abbr></td><td>89</td></tr>"
            "</table>"
        )
        soup = _soup(html)
        table = soup.find("table")
        items, _ = _parse_table_items(table)
        assert items[0].name == "Eli Manning"
        assert items[1].name == "Dak Prescott"
        assert items[2].name == "Tom Brady"


# ---------------------------------------------------------------------------
# _parse_ordered_list_items
# ---------------------------------------------------------------------------


class TestParseOrderedListItems:
    def test_basic_ol(self):
        soup = _soup("<ol><li>Jerry Rice</li><li>Emmitt Smith</li></ol>")
        items = _parse_ordered_list_items(soup)
        assert len(items) == 2
        assert items[0].name == "Jerry Rice"
        assert items[0].rank == 1
        assert items[1].name == "Emmitt Smith"
        assert items[1].rank == 2

    def test_no_ol_returns_empty(self):
        soup = _soup("<div>No list here</div>")
        items = _parse_ordered_list_items(soup)
        assert items == []

    def test_empty_ol_returns_empty(self):
        soup = _soup("<ol></ol>")
        items = _parse_ordered_list_items(soup)
        assert items == []


# ---------------------------------------------------------------------------
# parse_page_data  (integration)
# ---------------------------------------------------------------------------


class TestParsePageData:
    def _ranked_table_html(self) -> str:
        return _html(
            "<h1>NFL All-Time TD Leaders</h1>"
            "<table>"
            "<tr><th>Rank</th><th>Player</th><th>TDs</th></tr>"
            "<tr><td>1</td><td>Jerry Rice</td><td>208</td></tr>"
            "<tr><td>2</td><td>Emmitt Smith</td><td>175</td></tr>"
            "<tr><td>3</td><td>LaDainian Tomlinson</td><td>162</td></tr>"
            "</table>"
        )

    def test_parses_title(self):
        page = parse_page_data(
            "http://example.com", "nfl", "9", self._ranked_table_html()
        )
        assert page.title == "NFL All-Time TD Leaders"

    def test_parses_items(self):
        page = parse_page_data(
            "http://example.com", "nfl", "9", self._ranked_table_html()
        )
        assert page.item_count == 3
        assert page.items[0].name == "Jerry Rice"
        assert page.items[0].rank == 1

    def test_page_fields(self):
        page = parse_page_data(
            "http://example.com", "nfl", "9", self._ranked_table_html()
        )
        assert page.page_id == "9"
        assert page.url == "http://example.com"
        assert page.type == "list"
        assert page.clue_type == "rank"

    def test_answer_count_derived_from_top_n_title(self):
        html = _html(
            "<h1>Top 3 NFL All-Time TD Leaders</h1>"
            "<table>"
            "<tr><th>Rank</th><th>Player</th><th>TDs</th></tr>"
            "<tr><td>1</td><td>Jerry Rice</td><td>208</td></tr>"
            "</table>"
        )
        page = parse_page_data("http://example.com", "nfl", "9", html)
        assert page.answer_count == 3

    def test_answer_count_zero_when_no_top_n(self):
        page = parse_page_data(
            "http://example.com", "nfl", "9", self._ranked_table_html()
        )
        assert page.answer_count == 0

    def test_fallback_to_ol(self):
        html = _html(
            "<h1>Top Players</h1><ol><li>Alice</li><li>Bob</li><li>Carol</li></ol>"
        )
        page = parse_page_data("http://example.com", "test", "1", html)
        assert page.item_count == 3
        assert page.items[0].name == "Alice"

    def test_empty_page_returns_zero_items(self):
        html = _html("<p>No structured data here.</p>")
        page = parse_page_data("http://example.com", "test", "1", html)
        assert page.item_count == 0
        assert page.items == []


# ---------------------------------------------------------------------------
# _find_xlsx
# ---------------------------------------------------------------------------


class TestFindXlsx:
    def test_finds_existing_xlsx(self, tmp_path):
        book_dir = tmp_path / "nfl"
        book_dir.mkdir()
        xlsx = book_dir / "nfl_data.xlsx"
        xlsx.write_bytes(b"")
        result = _find_xlsx(str(tmp_path), "nfl")
        assert result == str(xlsx)

    def test_returns_none_when_no_book_dir(self, tmp_path):
        assert _find_xlsx(str(tmp_path), "nba") is None

    def test_returns_none_when_no_xlsx(self, tmp_path):
        book_dir = tmp_path / "nfl"
        book_dir.mkdir()
        (book_dir / "notes.txt").write_text("ignore")
        assert _find_xlsx(str(tmp_path), "nfl") is None


# ---------------------------------------------------------------------------
# _write_page_to_spreadsheet
# ---------------------------------------------------------------------------


class TestWritePageToSpreadsheet:
    def _make_xlsx(self, tmp_path, rows: Optional[list[dict]] = None) -> str:
        """Create a minimal Pages-sheet xlsx file."""
        xlsx_path = str(tmp_path / "test.xlsx")
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Pages"

        # Rows 1-3 are blank (header is row 4 per spreadsheet_store convention)
        # Row 4: headers
        headers = [
            "Page #",
            "Answer Key URL",
            "Title",
            "Description",
            "Type",
            "# Items / Clue Style",
        ]
        for col, h in enumerate(headers, start=1):
            ws.cell(row=4, column=col, value=h)

        # Data rows starting at row 5
        if rows:
            for row_idx, row_data in enumerate(rows, start=5):
                for col, h in enumerate(headers, start=1):
                    ws.cell(row=row_idx, column=col, value=row_data.get(h))

        wb.save(xlsx_path)
        return xlsx_path

    def test_writes_new_row(self, tmp_path):
        xlsx_path = self._make_xlsx(tmp_path)
        page = Page(
            page_id="9",
            url="http://example.com/9",
            title="TD Leaders",
            description="All-time TDs",
            type="list",
            clue_style="10 items",
        )
        _write_page_to_spreadsheet(page, xlsx_path)

        wb = openpyxl.load_workbook(xlsx_path)
        ws = wb["Pages"]
        # Row 5 should have our data
        assert ws.cell(row=5, column=1).value == 9
        assert ws.cell(row=5, column=2).value == "http://example.com/9"
        assert ws.cell(row=5, column=3).value == "TD Leaders"

    def test_overwrites_existing_row(self, tmp_path):
        xlsx_path = self._make_xlsx(
            tmp_path,
            rows=[
                {
                    "Page #": 9,
                    "Answer Key URL": "http://old.com",
                    "Title": "Old Title",
                    "Description": "",
                    "Type": "list",
                    "# Items / Clue Style": "",
                }
            ],
        )
        page = Page(
            page_id="9",
            url="http://new.com/9",
            title="New Title",
            description="Updated",
            type="list",
            clue_style="5 items",
        )
        _write_page_to_spreadsheet(page, xlsx_path)

        wb = openpyxl.load_workbook(xlsx_path)
        ws = wb["Pages"]
        # Row 5 should be updated, no duplicate row added
        assert ws.cell(row=5, column=2).value == "http://new.com/9"
        assert ws.cell(row=5, column=3).value == "New Title"
        # Row 6 should be empty (no duplicate)
        assert ws.cell(row=6, column=1).value is None

    def test_raises_when_no_pages_sheet(self, tmp_path):
        xlsx_path = str(tmp_path / "no_pages.xlsx")
        wb = openpyxl.Workbook()
        wb.active.title = "Sheet1"
        wb.save(xlsx_path)

        page = Page(page_id="1", url="http://example.com")
        with pytest.raises(ValueError, match="No 'Pages' sheet"):
            _write_page_to_spreadsheet(page, xlsx_path)
