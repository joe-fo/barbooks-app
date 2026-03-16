"""CLI tool to validate and simulate pulling trivia context from web.

Validates source data before spreadsheet entry.

Usage:
    python -m ingest --url <url> --book <book_id> --page <page_id> \\
        [--write] [--books-dir <path>] [--refresh]

Fetches the given URL (via page cache), parses structured data (ranked lists,
stats tables), previews the resulting Page record, and optionally writes it to
the spreadsheet.  Pass --refresh to force a live re-fetch and overwrite the cache.
"""

from __future__ import annotations

import argparse
import asyncio
import os
import re
import sys
from typing import Optional

import httpx
import openpyxl
from bs4 import BeautifulSoup, Tag

from app.domain.models import Page, PageItem

# ---------------------------------------------------------------------------
# Fetch
# ---------------------------------------------------------------------------


async def _fetch_html(url: str) -> str:
    """Fetch raw HTML from a URL."""
    headers = {
        "User-Agent": (
            "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36"
            " (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36"
        )
    }
    async with httpx.AsyncClient(follow_redirects=True) as client:
        response = await client.get(url, headers=headers, timeout=15.0)
        response.raise_for_status()
        return response.text


# ---------------------------------------------------------------------------
# Parse
# ---------------------------------------------------------------------------


def _extract_title(soup: BeautifulSoup) -> str:
    """Best-effort page title extraction."""
    # Try <h1> first, then <title>
    h1 = soup.find("h1")
    if h1 and h1.get_text(strip=True):
        return h1.get_text(strip=True)
    title_tag = soup.find("title")
    if title_tag:
        text = title_tag.get_text(strip=True)
        # Strip common suffixes like " - ESPN" or " | Wikipedia"
        text = re.split(r"\s*[-|]\s*", text)[0].strip()
        return text
    return ""


def _clean_name(name: str) -> str:
    """Strip appended abbreviated form from player names.

    Some pages embed both the full name and an abbreviated form in the same
    table cell, producing strings like 'Eli ManningE. Manning' or
    'Dak PrescottD. Prescott'.  This happens because ``get_text()`` concatenates
    text from nested ``<abbr>`` / ``<span>`` elements without a separator.

    Detects the pattern ``<...><LastWord><Initial>. <LastWord>`` and returns
    only the full-name prefix.
    """
    # Pattern: full name ends with a word that is immediately followed by
    # "X. <same_word>" — the abbreviated form appended without a space.
    m = re.match(r"^(.*\b(\w+))[A-Z]\.\s+\2\s*$", name)
    if m:
        return m.group(1).strip()
    return name


def _detect_stat_label(header_cells: list[str]) -> str:
    """Infer a stat label from table header cell text."""
    stat_keywords = [
        "td",
        "touchdown",
        "yards",
        "yds",
        "points",
        "pts",
        "wins",
        "losses",
        "stats",
        "receptions",
        "rec",
        "passing",
        "rushing",
        "score",
        "goals",
        "assists",
        "hits",
        "rbis",
        "era",
    ]
    for cell in header_cells:
        low = cell.lower().strip()
        for kw in stat_keywords:
            if kw in low:
                return cell.strip()
    return ""


def _parse_table_items(table: Tag) -> tuple[list[PageItem], str]:
    """Parse a <table> into PageItems.  Returns (items, stat_label)."""
    rows = table.find_all("tr")
    if not rows:
        return [], ""

    # Determine header row
    header_row = rows[0]
    header_cells = [th.get_text(strip=True) for th in header_row.find_all(["th", "td"])]
    stat_label = _detect_stat_label(header_cells)

    # Try to identify column positions
    name_col = stat_col = rank_col = None
    for i, cell in enumerate(header_cells):
        low = cell.lower()
        if any(k in low for k in ["name", "player", "team", "athlete"]):
            name_col = i
        elif any(k in low for k in ["rank", "#", "no.", "pos"]):
            rank_col = i
        elif stat_label and cell.strip() == stat_label:
            stat_col = i

    # Fall back: first col is rank, second is name, third is stat
    if rank_col is None and len(header_cells) >= 1:
        rank_col = 0
    if name_col is None and len(header_cells) >= 2:
        name_col = 1
    if stat_col is None and len(header_cells) >= 3:
        stat_col = 2
        if not stat_label:
            stat_label = header_cells[2] if len(header_cells) > 2 else ""

    items: list[PageItem] = []
    for row_idx, row in enumerate(rows[1:], start=1):
        cells = [td.get_text(strip=True) for td in row.find_all(["td", "th"])]
        if not cells or len(cells) < 2:
            continue

        # Attempt rank extraction
        rank_val: Optional[int] = None
        if rank_col is not None and rank_col < len(cells):
            raw = re.sub(r"[^\d]", "", cells[rank_col])
            if raw:
                rank_val = int(raw)
            else:
                rank_val = row_idx
        else:
            rank_val = row_idx

        name_val = _clean_name(
            cells[name_col] if name_col is not None and name_col < len(cells) else ""
        )
        stat_val = (
            cells[stat_col] if stat_col is not None and stat_col < len(cells) else ""
        )

        if not name_val:
            continue

        key = f"#{rank_val}" if rank_val else str(row_idx)
        items.append(
            PageItem(
                rank=rank_val,
                key=key,
                name=name_val,
                stat_value=stat_val,
                stat_label=stat_label,
            )
        )

    return items, stat_label


def _parse_ordered_list_items(soup: BeautifulSoup) -> list[PageItem]:
    """Fall back: parse <ol> lists for ranked items."""
    items: list[PageItem] = []
    for ol in soup.find_all("ol"):
        for i, li in enumerate(ol.find_all("li"), start=1):
            text = _clean_name(li.get_text(strip=True))
            if text:
                items.append(PageItem(rank=i, key=f"#{i}", name=text))
        if items:
            break
    return items


def parse_page_data(
    url: str,
    book_id: str,
    page_id: str,
    html: str,
) -> Page:
    """Parse raw HTML into a Page domain model."""
    soup = BeautifulSoup(html, "html.parser")
    title = _extract_title(soup)

    items: list[PageItem] = []
    stat_label = ""
    clue_type = ""

    # Try tables first
    tables = soup.find_all("table")
    for table in tables:
        candidate_items, candidate_stat = _parse_table_items(table)
        if len(candidate_items) >= 3:
            items = candidate_items
            stat_label = candidate_stat
            clue_type = "rank"
            break

    # Fall back to ordered lists
    if not items:
        items = _parse_ordered_list_items(soup)
        if items:
            clue_type = "rank"

    # Infer clue_type from stat label / title heuristics
    if not clue_type:
        low_title = title.lower()
        if any(k in low_title for k in ["year", "winner", "champion", "season"]):
            clue_type = "year"
        elif any(k in low_title for k in ["rank", "leader", "top", "best", "all-time"]):
            clue_type = "rank"
        else:
            clue_type = "rank"

    description = f"Parsed from {url}"

    return Page(
        page_id=page_id,
        url=url,
        title=title,
        description=description,
        type="list",
        clue_style=f"{len(items)} items" if items else "",
        clue_type=clue_type,
        item_count=len(items),
        stat_label=stat_label,
        items=items,
    )


# ---------------------------------------------------------------------------
# Display
# ---------------------------------------------------------------------------


def _preview_page(page: Page) -> None:
    """Print a structured preview of the parsed Page record."""
    print("\n" + "=" * 60)
    print("PARSED PAGE PREVIEW (dry run)")
    print("=" * 60)
    print(f"  page_id    : {page.page_id}")
    print(f"  url        : {page.url}")
    print(f"  title      : {page.title}")
    print(f"  description: {page.description}")
    print(f"  type       : {page.type}")
    print(f"  clue_type  : {page.clue_type}")
    print(f"  clue_style : {page.clue_style}")
    print(f"  stat_label : {page.stat_label}")
    print(f"  item_count : {page.item_count}")

    if page.items:
        print(f"\n  Items (showing first 10 of {len(page.items)}):")
        for item in page.items[:10]:
            stat_part = (
                f"  [{item.stat_value} {item.stat_label}]" if item.stat_value else ""
            )
            print(f"    {item.key:6s}  {item.name}{stat_part}")
        if len(page.items) > 10:
            print(f"    ... and {len(page.items) - 10} more")
    else:
        print("\n  Items: (none parsed)")

    print("=" * 60)


def _suggest_regex_patterns(page: Page) -> None:
    """Print suggested regex patterns for the deterministic short-circuit layer."""
    if not page.items:
        return

    print("\nSUGGESTED SHORT-CIRCUIT PATTERNS")
    print("-" * 60)
    print(f'  ("{page.items[0].stat_label or "stat"}" book_id, "{page.page_id}"): [')
    for item in page.items[:10]:
        escaped = re.escape(item.name)
        # Make whitespace flexible
        pattern = r"(?i)\b" + escaped.replace(r"\ ", r"\s+") + r"\b"
        stat_info = (
            f"{item.stat_value} {item.stat_label}".strip()
            if item.stat_value
            else "on this list"
        )
        rank_info = f"{item.key}" if item.rank else ""
        if item.rank:
            answer = f"Yes, {item.name} is {rank_info} with {stat_info}."
        else:
            answer = f"Yes, {item.name} is on this list."
        print(f'    (r"{pattern}", "{answer}"),')
    if len(page.items) > 10:
        print(f"    # ... {len(page.items) - 10} more items not shown")
    print("  ]")
    print("-" * 60)


# ---------------------------------------------------------------------------
# Write to spreadsheet
# ---------------------------------------------------------------------------


def _find_xlsx(books_dir: str, book_id: str) -> Optional[str]:
    """Find the xlsx file for a given book_id."""
    book_path = os.path.join(books_dir, book_id)
    if not os.path.isdir(book_path):
        return None
    for fname in os.listdir(book_path):
        if fname.endswith(".xlsx"):
            return os.path.join(book_path, fname)
    return None


def _write_page_to_spreadsheet(page: Page, xlsx_path: str) -> None:
    """Append or update the Page row in the spreadsheet's Pages sheet."""
    wb = openpyxl.load_workbook(xlsx_path)

    if "Pages" not in wb.sheetnames:
        raise ValueError(f"No 'Pages' sheet found in {xlsx_path}")

    ws = wb["Pages"]

    # Find the header row (row 4, 1-based openpyxl).
    # spreadsheet_store.py: pd.read_excel(..., header=3) → 0-indexed 3 → 1-indexed 4
    header_row_idx = 4
    headers = [ws.cell(row=header_row_idx, column=c).value for c in range(1, 20)]

    # Build column index map
    col_map: dict[str, int] = {}
    for i, h in enumerate(headers, start=1):
        if h is not None:
            col_map[str(h).strip()] = i

    # Check if page already exists
    page_num_col = col_map.get("Page #")
    if page_num_col is None:
        raise ValueError("Could not find 'Page #' column in Pages sheet")

    target_row: Optional[int] = None
    for row_idx in range(header_row_idx + 1, ws.max_row + 2):
        val = ws.cell(row=row_idx, column=page_num_col).value
        if val is None:
            target_row = row_idx  # first empty row
            break
        if str(int(val)) == page.page_id:
            target_row = row_idx  # overwrite existing
            break

    if target_row is None:
        target_row = ws.max_row + 1

    def _set(col_name: str, value) -> None:
        col = col_map.get(col_name)
        if col is not None:
            ws.cell(row=target_row, column=col, value=value)

    _set("Page #", int(page.page_id))
    _set("Answer Key URL", page.url)
    _set("Title", page.title)
    _set("Description", page.description)
    _set("Type", page.type)
    _set("# Items / Clue Style", page.clue_style)

    wb.save(xlsx_path)
    print(f"\n✓ Written page {page.page_id} to {xlsx_path}")


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------


def _build_page_from_cache_result(
    url: str,
    book_id: str,
    page_id: str,
    text: str,
    raw_items: list,
) -> Page:
    """Build a Page domain object from page_cache.get_or_fetch() results."""
    from app.domain.models import PageItem

    page_items: list[PageItem] = [
        PageItem(**item) if isinstance(item, dict) else item for item in raw_items
    ]
    stat_label = page_items[0].stat_label if page_items else ""
    clue_type = "rank" if any(item.rank for item in page_items) else ""
    title = next((line.strip() for line in text.splitlines() if line.strip()), "")

    return Page(
        page_id=page_id,
        url=url,
        title=title,
        description=f"Parsed from {url}",
        type="list",
        clue_style=f"{len(page_items)} items" if page_items else "",
        clue_type=clue_type,
        item_count=len(page_items),
        stat_label=stat_label,
        items=page_items,
    )


async def _run(args: argparse.Namespace) -> None:
    url = args.url
    book_id = args.book
    page_id = args.page
    books_dir = args.books_dir
    write = args.write
    show_patterns = args.patterns
    refresh = args.refresh

    from app.page_cache import get_or_fetch

    cache_label = " (force-refresh)" if refresh else ""
    print(f"Fetching {url}{cache_label} ...")
    try:
        text, raw_items = await get_or_fetch(book_id, page_id, url, refresh=refresh)
    except Exception as exc:
        print(f"ERROR: Failed to fetch URL: {exc}", file=sys.stderr)
        sys.exit(1)

    if text.startswith("Error"):
        print(f"ERROR: {text}", file=sys.stderr)
        sys.exit(1)

    print(f"Parsing data for book={book_id!r}, page={page_id!r} ...")
    page = _build_page_from_cache_result(url, book_id, page_id, text, raw_items)

    _preview_page(page)

    if show_patterns:
        _suggest_regex_patterns(page)

    # Dry-run: ask user to confirm write
    if not write:
        print("\n(Dry run — pass --write to commit this row to the spreadsheet)")
        return

    # Confirm
    try:
        answer = input("\nWrite this row to the spreadsheet? [y/N] ").strip().lower()
    except EOFError:
        answer = ""

    if answer not in ("y", "yes"):
        print("Aborted.")
        return

    xlsx_path = _find_xlsx(books_dir, book_id)
    if xlsx_path is None:
        print(
            f"ERROR: No xlsx file found for book '{book_id}' in {books_dir}",
            file=sys.stderr,
        )
        sys.exit(1)

    _write_page_to_spreadsheet(page, xlsx_path)


def main() -> None:
    parser = argparse.ArgumentParser(
        description=(
            "Validate and simulate pulling trivia context from web"
            " before spreadsheet entry."
        )
    )
    parser.add_argument("--url", required=True, help="Source URL for the trivia page")
    parser.add_argument("--book", required=True, help="Book ID (e.g. 'nfl')")
    parser.add_argument("--page", required=True, help="Page ID / number (e.g. '9')")
    parser.add_argument(
        "--books-dir",
        default="books",
        help="Path to the books directory (default: books/)",
    )
    parser.add_argument(
        "--write",
        action="store_true",
        help="After preview, prompt to write row to spreadsheet",
    )
    parser.add_argument(
        "--patterns",
        action="store_true",
        help="Also print suggested short-circuit regex patterns for the answer list",
    )
    parser.add_argument(
        "--refresh",
        action="store_true",
        help="Force a live re-fetch, bypassing the cache TTL",
    )

    args = parser.parse_args()
    asyncio.run(_run(args))


if __name__ == "__main__":
    main()
